# DDT - Data Driven Testing
# Reading data from excel:

# import openpyxl
#
# path="E:\SeleniumDemo\DDT_Data.xlsx"
#
# workbook=openpyxl.load_workbook(path)
# sheet=workbook.active                #sheet=workbook.get_sheet_by_name("sheet1")
#
# rows=sheet.max_row #5
# cols=sheet.max_column #4
#
# print(rows)
# print(cols)
#
# for r in range(1,rows+1):
#     for c in range(1,cols+1):
#         print(sheet.cell(row=r,column=c).value,end="  ")
#     print()

#Writing data to excel:

import openpyxl

path="E:\SeleniumDemo\DDT_Data_Write.xlsx"

workbook=openpyxl.load_workbook(path)
sheet=workbook.active

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r,column=c).value="welcome to python"

workbook.save(path)